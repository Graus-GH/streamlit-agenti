Perfetto — ecco il file **aggiornato** con `st.query_params` al posto di `st.experimental_get_query_params` (e tutto il resto identico alla versione iPad). Incolla tutto:

```python
import io
import re
from typing import List, Tuple

import numpy as np
import pandas as pd
import requests
import streamlit as st
from fpdf import FPDF

st.set_page_config(page_title="✨GRAUS Proposta Clienti", layout="wide")

# =========================
# CSS – desktop + iPad portrait responsive/touch
# =========================
st.markdown("""
<style>
/* Tabs radio */
div[data-testid="stRadio"] > div[role="radiogroup"]{
  display: inline-flex !important; gap: 8px !important; padding: 6px;
  border: 1px solid #cbd5e1; border-radius: 12px; background: #fff; margin: 6px 0 12px;
  flex-wrap: wrap;                        /* permetti a capo su schermi stretti */
}
div[data-testid="stRadio"] [role="radio"]{
  padding: 6px 12px; border: 1px solid transparent; border-radius: 10px;
  cursor: pointer; font-weight: 500; transition: background-color .15s, border-color .15s;
  display: inline-flex; align-items: center;
}
div[data-testid="stRadio"] [role="radio"][aria-checked="true"]{ background:#eaf2ff; border-color:#93c5fd; }
div[data-testid="stRadio"] [role="radio"]:hover{ background:#f1f5f9; }

/* Login card */
.login-card { border:1px solid #e2e8f0; border-radius:12px; padding:18px; background:#fff; }

/* Sidebar footer stick */
section[data-testid="stSidebar"] .block-container{ display:flex; flex-direction:column; height:100%; }
section[data-testid="stSidebar"] .block-container .stVerticalBlock:last-child{ margin-top:auto; border-top:1px solid #eee; padding-top:10px; }

/* Header logo right */
.header-right { text-align:right; }

/* Pulsanti full width + testo a capo */
div.stButton > button{
  width:100% !important; min-height:46px; white-space:normal; line-height:1.25; padding:10px 14px;
}

/* Data editor: aumenta font di cella e altezza riga dove possibile */
div[data-testid="stDataEditor"] *{ font-size:15px; }

/* ===== iPad Portrait: max 900px ===== */
@media (max-width: 900px){
  html, body { font-size: 18px; }
  .block-container{ padding-left: 10px; padding-right: 10px; }
  section[data-testid="stSidebar"]{ width: 300px !important; min-width: 300px !important; }

  /* Bottoni touch-friendly */
  div.stButton > button{ min-height: 56px; font-size: 18px; }

  /* Slider/inputs più facili da toccare */
  input[type="range"]{ height: 28px; }

  /* Checkbox più grandi dove possibile */
  input[type="checkbox"]{ width: 22px; height: 22px; }

  /* Data editor: padding */
  div[data-testid="stDataEditor"]{ padding-top: 4px; padding-bottom: 4px; }
}

/* Download buttons full width */
a[kind="download"]{ width:100%; display:inline-block; text-align:center; padding:12px 10px; }
</style>
""", unsafe_allow_html=True)

# =========================
# LOGIN – Utenti in chiaro (per test interni)
# =========================
USERS = {
    "merch":     {"name": "Merch",                "password": "Sterch"},
    "fdefazio":  {"name": "Francesco Defazio",    "password": "ciccio"},
    "kristin":   {"name": "Kristin",              "password": ">Smarter!"},
    "silvia":    {"name": "Silvia",               "password": ">Smarter!"},
    "angelo":    {"name": "Angelo",               "password": "NumeroUno"},
    "peppi":     {"name": "Peppi",                "password": ">Peppi25"},
    "luca":      {"name": "Luca",                 "password": ">Luca33"},
    "david":     {"name": "David",                "password": ">Dav!d"},
    "claudio":   {"name": "Claudio",              "password": "Claud!O"},
    "elsa":      {"name": "Elsa",                 "password": "ElsA!"},
    "barbara":   {"name": "Barbara",              "password": "BarbaBella"},
}

st.session_state.setdefault("authenticated", False)
st.session_state.setdefault("username", None)
st.session_state.setdefault("display_name", None)

def login_view():
    def norm(s: str) -> str:
        return (s or "").replace("\u00A0", " ").strip()

    st.title("🔐 Accesso richiesto")
    st.markdown('<div class="login-card">', unsafe_allow_html=True)
    with st.form("login_form", clear_on_submit=False):
        u_in = st.text_input("Username (es. merch) o Nome", placeholder="inserisci il Tuo nome utente...")
        p_in = st.text_input("Password", type="password")
        ok = st.form_submit_button("Accedi")

        if ok:
            u_key = norm(u_in).lower()
            pwd = norm(p_in)

            rec = USERS.get(u_key)
            if not rec:
                for uname, info in USERS.items():
                    if norm(info["name"]).lower() == u_key:
                        rec = info; u_key = uname; break

            if rec and pwd == rec["password"]:
                st.session_state.authenticated = True
                st.session_state.username = u_key
                st.session_state.display_name = rec["name"]
                st.success(f"Benvenuto, {rec['name']}!")
                st.rerun()
            else:
                st.error("Credenziali non valide. Controlla maiuscole/spazi.")
    st.markdown("</div>", unsafe_allow_html=True)

# =========================
# CONFIG – ORIGINE DATI
# =========================
SHEET_ID = "10BFJQTV1yL69cotE779zuR8vtG5NqKWOVH0Uv1AnGaw"; GID = "707323537"
FW_SHEET_ID = "1D4-zgwpAGiWDCpPwDVipAD7Nlpi4aesFwRRpud2W-rk"; FW_GID = "1549810072"

def gsheet_csv_export_url(sheet_id: str, gid: str) -> str:
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
def gsheet_csv_gviz_url(sheet_id: str, gid: str) -> str:
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&gid={gid}"

BASE_URLS = [gsheet_csv_export_url(SHEET_ID, GID), gsheet_csv_gviz_url(SHEET_ID, GID)]
FW_URLS   = [gsheet_csv_export_url(FW_SHEET_ID, FW_GID), gsheet_csv_gviz_url(FW_SHEET_ID, FW_GID)]

COL_MAP = {"codice": 0, "prodotto": 2, "categoria": 5, "tipologia": 6, "provenienza": 7, "prezzo": 8}
DISPLAY_COLUMNS = ["codice", "prodotto", "prezzo", "categoria", "tipologia", "provenienza"]
SEARCH_FIELDS = ["codice", "prodotto", "categoria", "tipologia", "provenienza"]

# =========================
# UTILS
# =========================
@st.cache_data(ttl=600)
def load_data(url_or_urls) -> pd.DataFrame:
    import time
    urls = [url_or_urls] if isinstance(url_or_urls, str) else list(url_or_urls)
    last_exc = None
    for url in urls:
        for attempt in range(3):
            try:
                r = requests.get(url, headers={"User-Agent": "Mozilla/5.0 (Streamlit App)"}, timeout=20)
                if r.status_code != 200:
                    last_exc = requests.HTTPError(f"HTTP {r.status_code} on {url}", response=r)
                    time.sleep(0.8*(attempt+1)); continue
                df_raw = pd.read_csv(io.BytesIO(r.content))
                df = pd.DataFrame({name: df_raw.iloc[:, idx] for name, idx in COL_MAP.items()})
                for c in ["prodotto","categoria","tipologia","provenienza"]:
                    df[c] = df[c].astype(str).fillna("").str.strip()
                def to_float(x):
                    if pd.isna(x): return np.nan
                    s = str(x).strip().replace("€","").replace(" ","")
                    s = s.replace(".", "").replace(",", ".") if "," in s and "." in s else s.replace(",", ".")
                    try: return float(s)
                    except: return np.nan
                df["prezzo"] = df["prezzo"].apply(to_float)
                def normalize_code(x: str) -> str:
                    s = re.sub("[^0-9]","",str(x)); s = s.lstrip("0") or "0"
                    if len(s)>2 and s.endswith("00"): s = s[:-2] or "0"
                    return s
                df["codice"] = df["codice"].astype(str).apply(normalize_code)
                df = df[(df["codice"]!="") & (df["prodotto"]!="")]
                df = df.sort_values(["prodotto","codice"], kind="stable").reset_index(drop=True)
                return df
            except requests.RequestException as e:
                last_exc = e; time.sleep(0.8*(attempt+1))
    raise last_exc if last_exc else RuntimeError("Errore nel download CSV")

def tokenize_query(q: str) -> List[str]:
    return [t for t in re.split(r"\s+", q.strip()) if t]

def row_matches(row: pd.Series, tokens: List[str], fields: List[str]) -> bool:
    haystack = " ".join(str(row[f]) for f in fields).lower()
    return all(t.lower() in haystack for t in tokens)

def adaptive_price_bounds(df: pd.DataFrame) -> Tuple[float, float]:
    if df["prezzo"].notna().any():
        mn = float(np.nanmin(df["prezzo"])); mx = float(np.nanmax(df["prezzo"]))
        if mn == mx: mx = mn + 0.01
        return max(0.0, round(mn, 2)), max(0.01, round(mx, 2))
    return 0.0, 0.01

def with_fw_prefix(df: pd.DataFrame) -> pd.DataFrame:
    d = df.copy()
    if "is_fw" in d.columns: d["prodotto"] = np.where(d["is_fw"], "⏳ " + d["prodotto"].astype(str), d["prodotto"])
    return d

# =========================
# EXPORT (Excel/PDF)
# =========================
def make_excel(df: pd.DataFrame) -> bytes:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Prodotti selezionati"
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.upper())
        cell.font = Font(name="Corbel", size=12, bold=True)
        cell.alignment = Alignment(horizontal="right" if col_name.lower()=="prezzo" else "left")
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = Font(name="Corbel", size=12)
            cell.alignment = Alignment(horizontal="right" if df.columns[c_idx-1].lower()=="prezzo" else "left")
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()

def make_pdf(df: pd.DataFrame) -> bytes:
    def pdf_safe(s: str) -> str:
        if s is None: return ""
        s = str(s).replace("⏳","[FW]"); return s.encode("latin-1","ignore").decode("latin-1")
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page(); pdf.set_font("Helvetica","B",14); pdf.cell(0,10,pdf_safe("Prodotti selezionati"), ln=1)
    pdf.set_font("Helvetica","B",10)
    headers = DISPLAY_COLUMNS; col_widths = [35, 120, 30, 40, 40, 40]
    for h, w in zip(headers, col_widths): pdf.cell(w, 8, pdf_safe(h.upper()), border=1)
    pdf.ln(8); pdf.set_font("Helvetica", size=9)
    for _, r in df.iterrows():
        cells = [str(r.get("codice","")), str(r.get("prodotto",""))[:200],
                 ("" if pd.isna(r.get("prezzo")) else f"{r.get('prezzo'):.2f}"),
                 str(r.get("categoria","")), str(r.get("tipologia","")), str(r.get("provenienza",""))]
        for c, w in zip(cells, col_widths):
            txt = pdf_safe(c.replace("\n"," ")); 
            if len(txt)>80: txt = txt[:77]+"..."
            pdf.cell(w, 6, txt, border=1)
        pdf.ln(6)
    out = pdf.output(dest="S")
    return bytes(out) if isinstance(out, (bytes, bytearray)) else out.encode("latin-1","ignore")

# =========================
# APP UI (solo dopo login)
# =========================
def run_app():
    # STATE
    if "basket" not in st.session_state:
        st.session_state.basket = pd.DataFrame(columns=DISPLAY_COLUMNS + ["is_fw"])
    for k in ["res_select_all_toggle","basket_select_all_toggle","reset_res_selection",
              "reset_basket_selection","flash","include_fw","active_tab","ipad_mode"]:
        default = "Ricerca" if k=="active_tab" else False
        st.session_state.setdefault(k, default)

    # Query param ?ipad=1  --> usa st.query_params
    qp = st.query_params
    if "ipad" in qp and str(qp["ipad"]).lower() in ("1","true","yes"):
        st.session_state.ipad_mode = True

    # DATA
    with st.spinner("Caricamento dati…"):
        try:
            df_base = load_data(BASE_URLS)
        except Exception as e:
            st.error("❌ Impossibile caricare il *listino base*. Verifica condivisione e ID/GID.")
            st.caption(f"Prova ad aprire in incognito: {BASE_URLS[0]}")
            st.exception(e); st.stop()
    df_base["is_fw"] = False

    # HEADER
    left, spacer, right = st.columns([6, 4, 2])
    with left:
        st.title("✨GRAUS Proposta Clienti")
    with right:
        st.markdown("<div class='header-right'>"
                    "<img src='https://res.cloudinary.com/dct4tiqsl/image/upload/v1754315051/LogoGraus_j7d5jo.png' width='120'/>"
                    "</div>", unsafe_allow_html=True)

    # ==== SIDEBAR ====
    with st.sidebar:
        sb_top = st.container(); sb_bottom = st.container()
    with sb_top:
        st.header("🔎 Ricerca")
        st.toggle("📱 Modalità iPad (verticale)", key="ipad_mode", help="Tap target più grandi e layout ottimizzato.")
        with st.form("search_form_sidebar", clear_on_submit=False):
            q = st.text_input(
                "Cerca su codice, prodotto, categoria, tipologia, provenienza",
                placeholder="Es. 'riesling alto adige 0,75'",
            )
            st.checkbox(
                "Includi FINE WINES (⏳)\nDisponibilità salvo conferma e consegna minimo 3 settimane.",
                value=st.session_state.include_fw, key="include_fw",
            )

            df_all = df_base.copy()
            if st.session_state.include_fw:
                try:
                    df_fw = load_data(FW_URLS); df_fw["is_fw"] = True
                    df_all = pd.concat([df_all, df_fw], ignore_index=True)
                except Exception as e:
                    st.warning("⚠️ Impossibile caricare il foglio Fine Wines."); st.caption(str(e))

            tokens = tokenize_query(q) if q else []
            mask_text = (df_all.apply(lambda r: row_matches(r, tokens, SEARCH_FIELDS), axis=1)
                         if tokens else pd.Series(True, index=df_all.index))
            df_after_text = df_all.loc[mask_text]

            dyn_min, dyn_max = adaptive_price_bounds(df_after_text)
            col_min, col_max = st.columns(2) if st.session_state.ipad_mode else st.columns([1,1])
            with col_min:
                min_price_input = st.number_input("Min", min_value=0.0, value=float(dyn_min), step=0.1, format="%.2f")
            with col_max:
                max_price_input = st.number_input("Max", min_value=0.01, value=float(dyn_max), step=0.1, format="%.2f")
            max_for_slider = max(min_price_input, max_price_input)
            price_range = st.slider("Range", min_value=0.0,
                                    max_value=max(0.01, round(max_for_slider, 2)),
                                    value=(float(min_price_input), float(max_price_input)),
                                    step=0.1, label_visibility="collapsed")
            submitted_sidebar = st.form_submit_button("Cerca")
            if submitted_sidebar: st.session_state.active_tab = "Ricerca"

    # Filtri applicati
    min_price = min(price_range[0], price_range[1])
    max_price = max(price_range[0], price_range[1])
    mask_price = df_after_text["prezzo"].fillna(0.0).between(min_price, max_price)
    df_res = df_after_text.loc[mask_price].reset_index(drop=True)

    # Tabs
    basket_len = len(st.session_state.basket)
    labels = ["Ricerca", f"Prodotti selezionati ({basket_len})"]
    index = 0 if st.session_state.active_tab == "Ricerca" else 1
    choice = st.radio("", options=labels, index=index, horizontal=True, label_visibility="collapsed")
    st.session_state.active_tab = "Ricerca" if choice == labels[0] else "Prodotti"

    # === RICERCA ===
    if st.session_state.active_tab == "Ricerca":
        st.caption(f"Risultati: {len(df_res)}")

        col_sel, col_add = st.columns(2)
        all_on = st.session_state.res_select_all_toggle and not st.session_state.reset_res_selection
        with col_sel:
            if st.button("Deseleziona tutti i risultati" if all_on else "Seleziona tutti i risultati", key="res_toggle_btn"):
                st.session_state.res_select_all_toggle = not all_on
                st.session_state.reset_res_selection = not st.session_state.res_select_all_toggle
                st.rerun()
        add_btn = col_add.button("➕ Aggiungi selezionati", type="primary", key="add_to_basket_btn")

        if st.session_state.flash:
            f = st.session_state.flash
            {"success": st.success, "info": st.info, "warning": st.warning, "error": st.error}.get(
                f.get("type","success"), st.success)(f.get("msg",""))
            if not f.get("shown", False): st.session_state.flash["shown"] = True
            else: st.session_state.flash = None

        default_sel = st.session_state.res_select_all_toggle and not st.session_state.reset_res_selection
        df_res_display = with_fw_prefix(df_res)[DISPLAY_COLUMNS].copy()
        df_res_display.insert(0, "sel", default_sel)

        if st.session_state.ipad_mode:
            colcfg = {
                "sel": st.column_config.CheckboxColumn(label="", width=42, help="Seleziona"),
                "codice": st.column_config.TextColumn(width=80),
                "prodotto": st.column_config.TextColumn(width=360),
                "prezzo": st.column_config.NumberColumn(format="€ %.2f", width=100),
                "categoria": st.column_config.TextColumn(width=120),
                "tipologia": st.column_config.TextColumn(width=120),
                "provenienza": st.column_config.TextColumn(width=120),
            }
        else:
            colcfg = {
                "sel": st.column_config.CheckboxColumn(label="", width=38, help="Seleziona riga"),
                "codice": st.column_config.TextColumn(width=80),
                "prodotto": st.column_config.TextColumn(width=420),
                "prezzo": st.column_config.NumberColumn(format="€ %.2f", width=90),
                "categoria": st.column_config.TextColumn(width=160),
                "tipologia": st.column_config.TextColumn(width=160),
                "provenienza": st.column_config.TextColumn(width=160),
            }

        edited_res = st.data_editor(
            df_res_display, hide_index=True, use_container_width=True, num_rows="fixed",
            column_config=colcfg, disabled=["codice","prodotto","prezzo","categoria","tipologia","provenienza"],
            key="res_editor",
        )

        if add_btn:
            selected_mask = edited_res["sel"].fillna(False)
            selected_codes = set(edited_res.loc[selected_mask, "codice"].tolist())
            if selected_codes:
                df_to_add = df_res[df_res["codice"].isin(selected_codes)]
                combined = pd.concat([st.session_state.basket, df_to_add], ignore_index=True)
                combined = combined.drop_duplicates(subset=["codice"]).reset_index(drop=True)
                st.session_state.basket = combined
                st.session_state.res_select_all_toggle = False
                st.session_state.reset_res_selection = True
                st.session_state.flash = {"type":"success","msg":f"Aggiunti {len(df_to_add)} articoli al paniere.","shown":False}
                st.session_state.active_tab = "Prodotti"; st.rerun()
            else:
                st.info("Seleziona almeno un articolo dalla griglia.")

    # === PANIERE ===
    if st.session_state.active_tab == "Prodotti":
        basket = st.session_state.basket.copy()

        # Bottoni: su iPad 2x2, altrimenti 1x4
        if st.session_state.ipad_mode:
            c1, c2 = st.columns(2)
            c3, c4 = st.columns(2)
            cols = [c1, c2, c3, c4]
        else:
            cols = st.columns(4)

        col_sel, col_rm, col_xls, col_pdf = cols

        all_on_b = st.session_state.basket_select_all_toggle and not st.session_state.reset_basket_selection
        if col_sel.button("Deseleziona tutto il paniere" if all_on_b else "Seleziona tutto il paniere"):
            st.session_state.basket_select_all_toggle = not all_on_b
            st.session_state.reset_basket_selection = not st.session_state.basket_select_all_toggle

        remove_btn = col_rm.button("🗑️ Rimuovi selezionati")

        basket_sorted = st.session_state.basket.sort_values(
            ["categoria","tipologia","provenienza","prodotto"], kind="stable"
        ).reset_index(drop=True)
        export_df = with_fw_prefix(basket_sorted)[DISPLAY_COLUMNS].copy()

        xbuf = make_excel(export_df)
        col_xls.download_button("⬇️ Esporta Excel", data=xbuf,
                                file_name="prodotti_selezionati.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        pbuf = make_pdf(export_df)
        col_pdf.download_button("⬇️ Crea PDF", data=pbuf,
                                file_name="prodotti_selezionati.pdf", mime="application/pdf")

        default_sel_b = st.session_state.basket_select_all_toggle and not st.session_state.reset_basket_selection
        basket_display = with_fw_prefix(basket)[DISPLAY_COLUMNS].copy()
        basket_display.insert(0, "rm", default_sel_b)

        if st.session_state.ipad_mode:
            colcfg_b = {
                "rm": st.column_config.CheckboxColumn(label="", width=42, help="Rimuovi"),
                "codice": st.column_config.TextColumn(width=80),
                "prodotto": st.column_config.TextColumn(width=360),
                "prezzo": st.column_config.NumberColumn(format="€ %.2f", width=100),
                "categoria": st.column_config.TextColumn(width=120),
                "tipologia": st.column_config.TextColumn(width=120),
                "provenienza": st.column_config.TextColumn(width=120),
            }
        else:
            colcfg_b = {
                "rm": st.column_config.CheckboxColumn(label="", width=38, help="Seleziona per rimuovere"),
                "codice": st.column_config.TextColumn(width=80),
                "prodotto": st.column_config.TextColumn(width=420),
                "prezzo": st.column_config.NumberColumn(format="€ %.2f", width=90),
                "categoria": st.column_config.TextColumn(width=160),
                "tipologia": st.column_config.TextColumn(width=160),
                "provenienza": st.column_config.TextColumn(width=160),
            }

        edited_basket = st.data_editor(
            basket_display, hide_index=True, use_container_width=True, num_rows="fixed",
            column_config=colcfg_b, disabled=["codice","prodotto","prezzo","categoria","tipologia","provenienza"],
            key="basket_editor",
        )

        if remove_btn:
            selected_mask_b = edited_basket["rm"].fillna(False)
            selected_codes_b = set(edited_basket.loc[selected_mask_b, "codice"].tolist())
            if selected_codes_b:
                st.session_state.basket = st.session_state.basket[~st.session_state.basket["codice"].isin(selected_codes_b)].reset_index(drop=True)
                st.session_state.basket_select_all_toggle = False
                st.session_state.reset_basket_selection = True
                st.session_state.flash = {"type":"success","msg":"Rimossi articoli selezionati.","shown":False}
                st.session_state.active_tab = "Prodotti"; st.rerun()
            else:
                st.info("Seleziona almeno un articolo dal paniere.")

    # ==== Sidebar footer (utente + logout)
    with sb_bottom:
        lcol, rcol = st.columns([3,1])
        with lcol: st.caption(f"👤 {st.session_state.display_name}")
        with rcol:
            if st.button("Logout"):
                for k in ["authenticated","username","display_name"]: st.session_state[k] = None if k!="authenticated" else False
                st.rerun()

# =========================
# ENTRY POINT
# =========================
if not st.session_state.authenticated:
    login_view()
else:
    run_app()
```

Se vuoi, posso aggiungere anche un parametro `?tab=prodotti` per aprire direttamente il paniere.
